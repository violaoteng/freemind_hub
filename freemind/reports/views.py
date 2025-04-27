from django.shortcuts import render
from django.http import HttpResponse, HttpResponseServerError
from django.contrib.auth.decorators import user_passes_test
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Q
from appointment.models import Appointment, TherapistAvailability
from users.models import Therapist, User
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from openpyxl import Workbook
from django.db.models.functions import TruncMonth
from django.utils.timezone import make_aware
from django.db.models import F, ExpressionWrapper, FloatField
from django.core.cache import cache
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl.utils import get_column_letter


def is_admin(user):
    return user.is_authenticated and user.role == 'admin'

@user_passes_test(is_admin)
def view_reports(request):
    try:
        # First try to get cached data
        cache_key = f'reports_data_{timezone.now().date()}'
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return render(request, 'reports/view_reports.html', cached_data)

        # Generate fresh data if cache is empty
        today = timezone.now().date()
        start_of_week = today - timedelta(days=today.weekday())  # Monday of current week
        start_of_month = today.replace(day=1)  # First day of current month

        # Get filter from URL parameters
        filter_param = request.GET.get('filter', 'all')

        # Base queryset with filtering
        appointments = Appointment.objects.all()
        if filter_param == 'today':
            appointments = appointments.filter(date__date=today)
        elif filter_param == 'week':
            appointments = appointments.filter(date__gte=start_of_week)
        elif filter_param == 'month':
            appointments = appointments.filter(date__gte=start_of_month)

        # Calculate status counts with percentages
        status_counts = Appointment.objects.values('status') \
                                .annotate(count=Count('status')) \
                                .order_by('-count')
        total_count = sum(item['count'] for item in status_counts)
        for item in status_counts:
            item['percentage'] = (item['count'] / total_count) * 100 if total_count > 0 else 0

        # Get therapist workload
        therapist_workload = Therapist.objects.annotate(
            total_appointments=Count('therapist_appointments'),
            completed=Count('therapist_appointments', 
                          filter=Q(therapist_appointments__status='Completed')),
            cancelled=Count('therapist_appointments', 
                          filter=Q(therapist_appointments__status='Cancelled')),
            completion_rate=ExpressionWrapper(
                F('completed') * 100.0 / F('total_appointments'),
                output_field=FloatField()
            ),
            no_show=Count('therapist_appointments', 
                        filter=Q(therapist_appointments__status='No Show'))
        ).order_by('-total_appointments').select_related('user')

        # Get availability stats
        availability_stats = {
            'total_slots': TherapistAvailability.objects.filter(is_active=True).count(),
            'booked_slots': Appointment.objects.filter(
                therapist__availabilities__is_active=True
            ).distinct().count(),
            'utilization_rate': round(
                (Appointment.objects.count() / 
                 TherapistAvailability.objects.filter(is_active=True).count()) * 100, 2
            ) if TherapistAvailability.objects.filter(is_active=True).count() else 0,
        }

        # Prepare context
        context = {
            'today': today,
            'total_appointments': appointments.count(),
            'today_appointments': Appointment.objects.filter(date__date=today).count(),
            'weekly_appointments': Appointment.objects.filter(date__gte=start_of_week).count(),
            'monthly_appointments': Appointment.objects.filter(date__gte=start_of_month).count(),
            'monthly_trends': [
                {'month': item['month'].strftime('%Y-%m'), 'count': item['count']}
                for item in Appointment.objects.annotate(month=TruncMonth('date'))
                    .values('month')
                    .annotate(count=Count('id'))
                    .order_by('month')
            ],
            'status_counts': status_counts,
            'therapist_workload': therapist_workload,
            'availability_stats': availability_stats,
            'all_therapists': Therapist.objects.all(),
            'all_statuses': Appointment.STATUS_CHOICES,
            'current_filter': filter_param,
            'report_date': timezone.now().strftime("%Y-%m-%d %H:%M")
        }

        # Cache the data for 1 hour (3600 seconds)
        try:
            cache.set(cache_key, context, timeout=3600)
        except Exception as cache_error:
            print(f"Cache error (non-critical): {str(cache_error)}")
            # Continue without caching if there's an error

        return render(request, 'reports/view_reports.html', context)

    except Exception as e:
        print(f"Error in view_reports: {str(e)}")
        # Return a proper error response
        return render(request, 'core/404.html', {'error': str(e)})

@user_passes_test(is_admin)
def download_report(request):
    try:
        source = request.GET.get('source', 'full_report')
        report_format = request.GET.get('format', 'pdf').lower()
        data = get_report_data(request, source=source)
        
        if report_format == 'excel':
            return generate_excel_report(data)
        else:
            return generate_pdf_report(data)
    except Exception as e:
        print(f"Error in download_report: {str(e)}")
        return HttpResponseServerError("An error occurred while generating the report. Please try again later.")
    
def get_report_data(request, source='full_report'):
    today = timezone.now().date()
    base_data = {
        'report_date': timezone.now().strftime('%Y-%m-%d %H:%M'),
        'period': {
            'today': today,
            'week_ago': today - timedelta(days=7),
            'month_ago': today - timedelta(days=30),
        },
        'source': source,
        'report_title': 'Comprehensive System Report' if source != 'recent_activities' else 'Recent Activities Report',
        'focus_period': 'All Time' if source != 'recent_activities' else 'Last 7 Days'
    }

    if source == 'recent_activities':
        base_data.update({
            'recent_appointments': Appointment.objects.filter(
                date__gte=today - timedelta(days=7)
            ).select_related('patient', 'therapist__user').order_by('-date')[:100],
            'recent_signups': User.objects.filter(
                date_joined__gte=today - timedelta(days=7)
            ).order_by('-date_joined')[:50],
        })
    else:
        base_data.update({
            'appointments': {
                'total': Appointment.objects.count(),
                'today': Appointment.objects.filter(date__date=today).count(),
                'week': Appointment.objects.filter(date__gte=today - timedelta(days=7)).count(),
                'month': Appointment.objects.filter(date__gte=today - timedelta(days=30)).count(),
            },
            'status_counts': Appointment.objects.values('status')
                                .annotate(count=Count('status'))
                                .order_by('-count'),
            'therapist_workload': Therapist.objects.annotate(
                total_appointments=Count('therapist_appointments'),
                completed=Count('therapist_appointments', filter=Q(therapist_appointments__status='Completed')),
                cancelled=Count('therapist_appointments', filter=Q(therapist_appointments__status='Cancelled')),
                completion_rate=ExpressionWrapper(
                    F('completed') * 100.0 / F('total_appointments'),
                    output_field=FloatField()
                ),
                no_show=Count('therapist_appointments', filter=Q(therapist_appointments__status='No Show'))
            ).order_by('-total_appointments'),
            'availability_stats': {
                'total_slots': TherapistAvailability.objects.filter(is_active=True).count(),
                'booked_slots': Appointment.objects.filter(
                    therapist__availabilities__is_active=True
                ).distinct().count(),
                'utilization_rate': round(
                    (Appointment.objects.count() / TherapistAvailability.objects.filter(is_active=True).count()) * 100, 2
                ) if TherapistAvailability.objects.filter(is_active=True).count() else 0,
            },
        })
    return base_data
    
def generate_pdf_report(data):
    buffer = BytesIO()
    try:
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=20,
            rightMargin=20,
            topMargin=40,
            bottomMargin=40
        )
        styles = getSampleStyleSheet()
        elements = []

        # Title Section
        elements.append(Paragraph(data['report_title'], styles['Title']))
        elements.append(Paragraph(f"Generated: {data['report_date']}", styles['Normal']))
        elements.append(Paragraph(f"Period: {data['focus_period']}", styles['Normal']))
        elements.append(Spacer(1, 24))

        if data['source'] == 'recent_activities':
            # Recent Appointments Table
            table_data = [['Date', 'Patient', 'Therapist', 'Status']]
            table_data.extend([
                [
                    appt.date.strftime('%Y-%m-%d %H:%M'),
                    appt.patient.get_full_name(),
                    appt.therapist.user.get_full_name(),
                    appt.status
                ] for appt in data['recent_appointments']
            ])
            
            # Create and style table
            table = Table(table_data, colWidths=[120, 150, 150, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#D9E1F2')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('WORDWRAP', (0,0), (-1,-1), True)
            ]))
            elements.append(table)
        else:
            # Full Report Tables
            # Summary Table
            summary_data = [
                ['Metric', 'Count'],
                ['Total Appointments', data['appointments']['total']],
                ['Today', data['appointments']['today']],
                ['This Week', data['appointments']['week']],
                ['This Month', data['appointments']['month']]
            ]
            
            # Status Breakdown Table
            status_data = [['Status', 'Count', 'Percentage']]
            status_data.extend([
                [item['status'], item['count'], f"{item.get('percentage', 0):.1f}%"] 
                for item in data.get('status_counts', [])
            ])
            
            # Therapist Workload Table
            therapist_data = [['Therapist', 'Total', 'Completed', 'Cancelled', 'No Show', 'Rate']]
            therapist_data.extend([
                [therapist.user.get_full_name(), 
                 therapist.total_appointments,
                 therapist.completed,
                 therapist.cancelled,
                 therapist.no_show,
                 f"{therapist.completion_rate:.1f}%" if therapist.completion_rate else "N/A"]
                for therapist in data.get('therapist_workload', [])
            ])
            
            # Availability Stats Table
            availability_data = [
                ['Total Slots', data.get('availability_stats', {}).get('total_slots', 'N/A')],
                ['Booked Slots', data.get('availability_stats', {}).get('booked_slots', 'N/A')],
                ['Utilization Rate', f"{data.get('availability_stats', {}).get('utilization_rate', 0)}%"]
            ]
            
            # Create all tables
            tables = [
                ('Summary', summary_data, [150, 80]),
                ('Status Breakdown', status_data, [100, 80, 80]),
                ('Therapist Workload', therapist_data, [150, 60, 60, 60, 60, 60]),
                ('Availability Statistics', availability_data, [150, 80])
            ]
            
            for title, t_data, col_widths in tables:
                elements.append(Paragraph(title, styles['Heading2']))
                table = Table(t_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,0), 10),
                    ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#D9E1F2')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('WORDWRAP', (0,0), (-1,-1), True)
                ]))
                elements.append(table)
                elements.append(Spacer(1, 12))

        # Build the document and return response (this was outside the if-else block)
        doc.build(elements)
        pdf_data = buffer.getvalue()
        response = HttpResponse(
            content_type='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{data["source"]}_report.pdf"',
                'Content-Length': str(len(pdf_data))
            }
        )
        response.write(pdf_data)
        return response
    finally:
        buffer.close()
        
    
def generate_excel_report(data):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{data["source"]}_report.xlsx"'
        }
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Summary"
    
    # Header
    ws.append([data['report_title']])
    ws.append([f"Generated: {data['report_date']}"])
    ws.append([f"Period: {data['focus_period']}"])
    ws.append([])
    
    if data['source'] == 'recent_activities':
        # Recent Activities
        ws.append(["Recent Appointments"])
        ws.append(["Date", "Patient", "Therapist", "Status"])
        for appt in data['recent_appointments']:
            ws.append([
                appt.date.strftime('%Y-%m-%d %H:%M'),
                appt.patient.get_full_name(),
                appt.therapist.user.get_full_name(),
                appt.status
            ])
    else:
        # Full Report
        ws.append(["Metric", "Count"])
        ws.append(["Total Appointments", data['appointments']['total']])
        ws.append(["Today", data['appointments']['today']])
        ws.append(["This Week", data['appointments']['week']])
        ws.append(["This Month", data['appointments']['month']])
        
        ws.append([])
        ws.append(["Status Breakdown"])
        ws.append(["Status", "Count"])
        for item in data['status_counts']:
            ws.append([item['status'], item['count']])
        
        ws.append([])
        ws.append(["Therapist Workload"])
        ws.append(["Therapist", "Total", "Completed", "Cancelled", "Completion Rate"])
        for therapist in data['therapist_workload']:
            ws.append([
                therapist.user.get_full_name(),
                therapist.total_appointments,
                therapist.completed,
                therapist.cancelled,
                f"{therapist.completion_rate:.1f}%" if therapist.completion_rate else "N/A"
            ])
    
    # Auto-adjust columns
    for col in ws.columns:
        max_length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = (max_length + 2) * 1.2
    
    wb.save(response)
    return response